import streamlit as st
import anthropic
import openpyxl
import json
import io
import requests
from bs4 import BeautifulSoup
import re

st.set_page_config(
    page_title="OSB Lead Generation Ajanı",
    page_icon="🏭",
    layout="wide"
)

st.markdown("""
<style>
.main-title { font-size: 1.8rem; font-weight: 800; color: #1F4E79; }
.step-badge { display: inline-block; background: #1F4E79; color: white; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 700; margin-bottom: 8px; }
.osb-card { background: #f0f4ff; border: 2px solid #d0d8f0; border-radius: 10px; padding: 16px; margin-bottom: 8px; }
.stat-box { background: #f8f9fc; border-radius: 8px; padding: 14px; text-align: center; border: 1px solid #e0e4f0; }
.stat-val { font-size: 26px; font-weight: 800; color: #1F4E79; }
.stat-label { font-size: 11px; color: #666; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
}

# ============ SESSION STATE ============
for key, val in [('adim',1),('sehir',''),('osblar',[]),('secili_osb',None),('tum_firmalar',[]),('gosterilen_ids',set())]:
    if key not in st.session_state:
        st.session_state[key] = val


def extract_contact(text):
    email = re.search(r'[\w\.\-]+@[\w\.\-]+\.\w{2,}', text)
    tel = re.search(r'(\+90[\s\-]?|0)[\s\-\(\)]?([2-5]\d{2})[\s\-\(\)]?(\d{3})[\s\-]?(\d{2})[\s\-]?(\d{2})', text)
    web = re.search(r'https?://[^\s<>"]+|www\.[^\s<>"]+', text)
    # WhatsApp: wa.me/90XXXXXXXXXX veya "whatsapp" yakınında geçen numara
    wa = re.search(r'wa\.me/(\d+)', text)
    return {
        'email': email.group() if email else None,
        'telefon': tel.group(0).strip() if tel else None,
        'website': web.group() if web else None,
        'whatsapp': wa.group(1) if wa else None,
    }


def iletisim_sayfasindan_cek(website):
    """Firmanın sayfalarından gerçek telefon, email ve whatsapp çeker"""
    if not website:
        return {}
    if not website.startswith('http'):
        website = 'https://' + website

    iletisim_yollari = ['', '/iletisim', '/contact', '/iletisim.html',
                        '/contact.html', '/hakkimizda', '/about', '/tr/iletisim']
    telefon = None
    email = None
    whatsapp = None

    for yol in iletisim_yollari:
        try:
            url = website.rstrip('/') + yol
            r = requests.get(url, headers=HEADERS, timeout=6)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')
            text = soup.get_text(separator=' ')
            html = r.text

            # Telefon - Türkiye formatı
            if not telefon:
                tel_match = re.search(
                    r'(\+90[\s\-]?|0)[\s\-\(\)]?([2-5]\d{2})[\s\-\(\)]?(\d{3})[\s\-]?(\d{2})[\s\-]?(\d{2})',
                    text
                )
                if tel_match:
                    telefon = re.sub(r'\s+', ' ', tel_match.group(0)).strip()

            # Email
            if not email:
                email_match = re.search(
                    r'[\w\.\-]+@[\w\.\-]+\.(com|net|org|com\.tr|org\.tr|net\.tr|edu\.tr)',
                    text
                )
                if email_match:
                    email = email_match.group(0)

            # WhatsApp - wa.me linki
            if not whatsapp:
                wa_match = re.search(r'wa\.me/(\d+)', html)
                if wa_match:
                    num = wa_match.group(1)
                    # Türkiye kodu varsa düzenle
                    if num.startswith('90'):
                        whatsapp = '+' + num
                    else:
                        whatsapp = '+90' + num

            # WhatsApp - href="whatsapp://..." veya data-phone ile
            if not whatsapp:
                wa_href = re.search(r'whatsapp://send\?phone=([\d\+]+)', html)
                if wa_href:
                    whatsapp = wa_href.group(1)

            # WhatsApp - "whatsapp" kelimesinin yanındaki numara
            if not whatsapp:
                wa_text = re.search(
                    r'(?i)whatsapp[^\d]{0,20}(\+?90[\s\-]?[\d\s\-]{10,14})',
                    text
                )
                if wa_text:
                    whatsapp = re.sub(r'\s+', '', wa_text.group(1))

            if telefon and email and whatsapp:
                break

        except:
            continue

    result = {}
    if telefon:
        result['telefon'] = telefon
    if email:
        result['email'] = email
    if whatsapp:
        result['whatsapp'] = whatsapp
    return result


def firmalar_cek_osb_sitesi(osb_website, osb_ad, limit=100):
    firmalar = []
    yollar = ['/firmalar', '/firma-listesi', '/firmalar/', '/uye-firmalar',
              '/rehber', '/firma-rehberi', '/uyeler', '/firmalar.html',
              '/tr/firmalar', '/firma_listesi']

    for yol in yollar:
        try:
            url = osb_website.rstrip('/') + yol
            r = requests.get(url, headers=HEADERS, timeout=8)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, 'html.parser')

            for table in soup.find_all('table'):
                for row in table.find_all('tr')[1:]:
                    cols = row.find_all(['td', 'th'])
                    if not cols:
                        continue
                    ad = cols[0].get_text(strip=True)
                    if len(ad) < 4:
                        continue
                    firma = {'ad': ad, 'osb': osb_ad}
                    row_text = row.get_text()
                    c = extract_contact(row_text)
                    firma.update({k: v for k, v in c.items() if v})
                    if ad not in [f['ad'] for f in firmalar]:
                        firmalar.append(firma)

            if not firmalar:
                for el in soup.find_all(['li', 'div', 'tr']):
                    text = el.get_text(strip=True)
                    if ('A.Ş' in text or 'LTD' in text.upper() or 'SAN.' in text.upper()) and 5 < len(text) < 400:
                        ad = text[:100].split('\n')[0].strip()
                        if ad and ad not in [f['ad'] for f in firmalar]:
                            firma = {'ad': ad, 'osb': osb_ad}
                            c = extract_contact(text)
                            firma.update({k: v for k, v in c.items() if v})
                            firmalar.append(firma)

            if firmalar:
                break
        except:
            continue

    return firmalar[:limit]


def firmalar_cek_ergene(osb_ad, sehir, limit=100):
    osb_siteleri = {
        'ergene': 'https://www.ergene1osb.org/firmarehberi/',
        'tuzla': 'https://www.itosb.org.tr/firmalar',
        'baskent': 'https://www.baskentosb.org/tr/firma-listesi/',
        'uludag': 'https://uludagosb.tr/firmalar/',
        'turgutlu': 'https://www.turgutluosb.org.tr/firmalar/',
    }
    firmalar = []
    for anahtar, url in osb_siteleri.items():
        if anahtar.lower() in osb_ad.lower() or anahtar.lower() in sehir.lower():
            try:
                r = requests.get(url, headers=HEADERS, timeout=10)
                soup = BeautifulSoup(r.text, 'html.parser')
                for el in soup.find_all(['td', 'li', 'div']):
                    text = el.get_text(strip=True)
                    if ('A.Ş' in text or 'LTD' in text.upper()) and 5 < len(text) < 300:
                        ad = text.split('\n')[0][:100].strip()
                        if ad and ad not in [f['ad'] for f in firmalar]:
                            firma = {'ad': ad, 'osb': osb_ad, 'sehir': sehir}
                            c = extract_contact(text)
                            firma.update({k: v for k, v in c.items() if v})
                            firmalar.append(firma)
            except:
                pass
    return firmalar[:limit]


def claude_osb_listesi(sehir, api_key):
    client = anthropic.Anthropic(api_key=api_key)
    r = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        messages=[{"role": "user", "content": f"""
{sehir} iline bağlı TÜM Organize Sanayi Bölgelerini listele. İlçelerdekiler dahil.
Her OSB için resmi website adresini yaz (örn: https://www.izmairosb.org.tr).

Sadece JSON döndür:
{{
  "osblar": [
    {{
      "id": 1,
      "ad": "OSB tam adı",
      "website": "https://... veya null",
      "ilce": "ilçe adı",
      "firma_sayisi": 150
    }}
  ]
}}
"""}]
    )
    try:
        text = r.content[0].text
        clean = text[text.find('{'):text.rfind('}')+1]
        data = json.loads(clean)
        osblar = data.get('osblar', [])
        for i, o in enumerate(osblar):
            o['id'] = i
        return osblar
    except:
        return []


def claude_firma_listesi(sehir, osb_ad, haric_listesi, adet, api_key):
    """
    Claude'dan SADECE firma adı + website ister.
    haric_listesi: daha önce çekilen firma adları — bunları tekrar verme.
    """
    client = anthropic.Anthropic(api_key=api_key)

    # Claude'a gönderilecek "hariç" listesini kısalt (token tasarrufu)
    haric_str = json.dumps(haric_listesi[:100], ensure_ascii=False)

    r = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4000,
        messages=[{"role": "user", "content": f"""
{sehir} şehrindeki {osb_ad} içinde faaliyet gösteren firmaları listele.
ZORUNLU KURAL: Aşağıdaki firmaları KESİNLİKLE tekrar verme, bunlar zaten listemde var:
{haric_str}

{adet} FARKLI ve YENİ firma ver. Telefon veya email YAZMA.
Sadece gerçek firma adları ve varsa websiteleri.

Sadece JSON döndür, başka hiçbir şey yazma:
{{"firmalar":[{{"ad":"Firma Adı A.Ş.","website":"www.firma.com.tr","sektor":"Tekstil"}}]}}
"""}]
    )
    try:
        text = r.content[0].text
        clean = text[text.find('{'):text.rfind('}')+1]
        firmalar = json.loads(clean).get('firmalar', [])
        # Ekstra güvenlik: hariç listesindeki firmalar gelirse çıkar
        haric_set = set(h.lower().strip() for h in haric_listesi)
        firmalar = [f for f in firmalar if f.get('ad','').lower().strip() not in haric_set]
        return firmalar
    except:
        return []


# ============ SIDEBAR ============
with st.sidebar:
    st.markdown("### ⚙️ Ayarlar")
    api_key = st.text_input("Anthropic API Key", type="password", placeholder="sk-ant-...")
    st.divider()
    st.markdown("### 📍 Adımlar")
    adimlar = ["1️⃣ Şehir Gir", "2️⃣ OSB Seç", "3️⃣ Firma Listesi", "4️⃣ Excel İndir"]
    for i, a in enumerate(adimlar, 1):
        if i < st.session_state.adim:
            st.markdown(f"✅ ~~{a}~~")
        elif i == st.session_state.adim:
            st.markdown(f"**▶ {a}**")
        else:
            st.markdown(f"⏳ {a}")
    st.divider()
    if st.button("🔄 Sıfırla", use_container_width=True):
        for key in ['adim','sehir','osblar','secili_osb','tum_firmalar','gosterilen_ids']:
            if key in st.session_state:
                del st.session_state[key]
        st.rerun()

# ============ BAŞLIK ============
st.markdown('<div class="main-title">🏭 OSB Lead Generation Ajanı</div>', unsafe_allow_html=True)
st.markdown("---")

# ============ ADIM 1 ============
if st.session_state.adim == 1:
    st.markdown('<div class="step-badge">ADIM 1 / 4</div>', unsafe_allow_html=True)
    st.markdown("### 🏙️ Hangi şehrin OSB'lerini aratalım?")

    col1, col2 = st.columns([3, 1])
    with col1:
        sehir_input = st.text_input("Şehir adı", placeholder="örn: Gaziantep, Bursa, İzmir...", label_visibility="collapsed")
    with col2:
        ara_btn = st.button("🔍 OSB'leri Bul", type="primary", use_container_width=True)

    if ara_btn:
        if not api_key:
            st.error("⚠️ Sol panelden API Key gir!")
            st.stop()
        if not sehir_input:
            st.error("⚠️ Şehir adı gir!")
            st.stop()

        with st.spinner(f"{sehir_input} OSB'leri aranıyor..."):
            osblar = claude_osb_listesi(sehir_input, api_key)
            if not osblar:
                osblar = [{"id": 0, "ad": f"{sehir_input} OSB", "website": None, "ilce": "Merkez", "firma_sayisi": 200}]

        st.session_state.sehir = sehir_input
        st.session_state.osblar = osblar
        st.session_state.adim = 2
        st.rerun()

# ============ ADIM 2 ============
elif st.session_state.adim == 2:
    st.markdown('<div class="step-badge">ADIM 2 / 4</div>', unsafe_allow_html=True)
    st.markdown(f"### 🏭 {st.session_state.sehir} — OSB Seç")
    st.markdown(f"*{len(st.session_state.osblar)} OSB bulundu:*")

    for osb in st.session_state.osblar:
        col1, col2 = st.columns([4, 1])
        with col1:
            web = osb.get('website', '')
            web_html = f" | 🌐 {web}" if web else ''
            st.markdown(f"""
            <div class="osb-card">
                <strong style="color:#1F4E79">{osb['ad']}</strong><br>
                <span style="font-size:12px;color:#666">📍 {osb.get('ilce','-')} | 🏭 ~{osb.get('firma_sayisi','?')} firma{web_html}</span>
            </div>""", unsafe_allow_html=True)
        with col2:
            if st.button("Seç →", key=f"osb_{osb['id']}", use_container_width=True):
                st.session_state.secili_osb = osb
                st.session_state.tum_firmalar = []
                st.session_state.gosterilen_ids = set()
                st.session_state.adim = 3
                st.rerun()

    if st.button("← Geri"):
        st.session_state.adim = 1
        st.rerun()

# ============ ADIM 3 ============
elif st.session_state.adim == 3:
    st.markdown('<div class="step-badge">ADIM 3 / 4</div>', unsafe_allow_html=True)
    osb = st.session_state.secili_osb
    toplam = osb.get('firma_sayisi') or 200

    st.markdown(f"### 🔍 {osb['ad']} — Firma Listesi")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div class="stat-box"><div class="stat-val">{toplam}</div><div class="stat-label">Tahmini Firma</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="stat-box"><div class="stat-val">{len(st.session_state.tum_firmalar)}</div><div class="stat-label">Listelenen</div></div>', unsafe_allow_html=True)
    with c3:
        kalan = max(0, toplam - len(st.session_state.tum_firmalar))
        st.markdown(f'<div class="stat-box"><div class="stat-val">{kalan}</div><div class="stat-label">Kalan</div></div>', unsafe_allow_html=True)

    st.markdown("")
    adet = st.slider("Kaç firma listeleyelim?", 10, 1000, 50, 10)

    if st.button(f"📋 {adet} Firma Listele", type="primary", use_container_width=True):
        yeni_firmalar = []
        osb_website = osb.get('website', '')

        # Mevcut firma adları — bunları tekrar çekme
        mevcut_adlar = list(st.session_state.gosterilen_ids)

        with st.spinner("Veriler çekiliyor..."):

            # 1. OSB'nin kendi sitesinden dene
            if osb_website:
                with st.status(f"🌐 {osb_website} taranıyor..."):
                    tum = firmalar_cek_osb_sitesi(osb_website, osb['ad'], limit=adet + len(mevcut_adlar))
                    # Daha önce çekilenleri çıkar
                    mevcut_set = set(a.lower().strip() for a in mevcut_adlar)
                    yeni_firmalar = [f for f in tum if f['ad'].lower().strip() not in mevcut_set]
                    yeni_firmalar = yeni_firmalar[:adet]
                    st.write(f"→ {len(yeni_firmalar)} yeni firma bulundu")

            # 2. Bilinen OSB sitelerini dene
            if not yeni_firmalar:
                with st.status("🔍 Bilinen OSB siteleri taranıyor..."):
                    tum = firmalar_cek_ergene(osb['ad'], st.session_state.sehir, limit=adet + len(mevcut_adlar))
                    mevcut_set = set(a.lower().strip() for a in mevcut_adlar)
                    yeni_firmalar = [f for f in tum if f['ad'].lower().strip() not in mevcut_set]
                    yeni_firmalar = yeni_firmalar[:adet]
                    st.write(f"→ {len(yeni_firmalar)} yeni firma bulundu")

            # 3. Claude ile liste oluştur
            if not yeni_firmalar and api_key:
                with st.status("🤖 Claude ile yeni firmalar oluşturuluyor..."):
                    yeni_firmalar = claude_firma_listesi(
                        st.session_state.sehir,
                        osb['ad'],
                        mevcut_adlar,   # ← mevcut tüm firma adlarını gönder
                        adet,
                        api_key
                    )
                    st.write(f"→ {len(yeni_firmalar)} yeni firma listelendi")

            # 4. Her firmanın sitesinden iletişim bilgisi çek
            if yeni_firmalar:
                with st.status("📞 İletişim bilgileri çekiliyor (telefon, email, WhatsApp)..."):
                    for i, f in enumerate(yeni_firmalar):
                        web = f.get('website', '')
                        if web:
                            st.write(f"🔍 [{i+1}/{len(yeni_firmalar)}] {f['ad']}...")
                            iletisim = iletisim_sayfasindan_cek(web)
                            f.update(iletisim)
                    st.write("✅ Tamamlandı")

        # Tekrar kontrolü yaparak ekle
        eklenen = 0
        for f in yeni_firmalar:
            ad = f.get('ad', '').strip()
            if ad and ad not in st.session_state.gosterilen_ids:
                f.setdefault('osb', osb['ad'])
                f.setdefault('sehir', st.session_state.sehir)
                st.session_state.tum_firmalar.append(f)
                st.session_state.gosterilen_ids.add(ad)
                eklenen += 1

        if eklenen:
            st.success(f"✅ {eklenen} yeni firma eklendi! Toplam: {len(st.session_state.tum_firmalar)}")
        else:
            st.warning("Yeni firma bulunamadı.")
        st.rerun()

    # Firma listesi göster
    if st.session_state.tum_firmalar:
        st.markdown("---")
        st.markdown(f"**📋 {len(st.session_state.tum_firmalar)} Firma**")
        for f in st.session_state.tum_firmalar:
            email = f.get('email', '')
            tel = f.get('telefon', '')
            wa = f.get('whatsapp', '')
            if email and tel:
                renk, durum = "#E1F5EE", "✅"
            elif email or tel or wa:
                renk, durum = "#FAEEDA", "⚠️"
            else:
                renk, durum = "#FCEBEB", "❌"
            wa_html = f" &nbsp;|&nbsp; 💬 {wa}" if wa else ""
            st.markdown(f"""
            <div style="background:{renk};border-radius:6px;padding:10px 14px;margin-bottom:6px;font-size:13px;">
                {durum} <strong>{f['ad']}</strong> &nbsp;|&nbsp; 📧 {email or '—'} &nbsp;|&nbsp; 📞 {tel or '—'}{wa_html} &nbsp;|&nbsp; 🌐 {f.get('website','—')}
            </div>""", unsafe_allow_html=True)

        col_x, col_y = st.columns(2)
        with col_x:
            if st.button("⬇️ Excel'e Aktar", type="primary", use_container_width=True):
                st.session_state.adim = 4
                st.rerun()
        with col_y:
            if st.button("← OSB Seçimine Dön", use_container_width=True):
                st.session_state.adim = 2
                st.rerun()

# ============ ADIM 4 ============
elif st.session_state.adim == 4:
    st.markdown('<div class="step-badge">ADIM 4 / 4</div>', unsafe_allow_html=True)
    firmalar = st.session_state.tum_firmalar

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{st.session_state.sehir} OSB"

    basliklar = ["#", "Firma Adı", "OSB", "Sektör", "Web Sitesi", "Email", "Telefon", "WhatsApp", "Durum"]
    for col, b in enumerate(basliklar, 1):
        h = ws.cell(row=1, column=col, value=b)
        h.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        h.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        h.alignment = openpyxl.styles.Alignment(horizontal="center")

    tam = kismi = eksik_count = 0
    for row, f in enumerate(firmalar, 2):
        email = f.get('email', '')
        tel = f.get('telefon', '')
        wa = f.get('whatsapp', '')
        if email and tel:
            durum, renk, tam = "✅ Tam", "E1F5EE", tam + 1
        elif email or tel or wa:
            durum, renk, kismi = "⚠️ Kısmi", "FAEEDA", kismi + 1
        else:
            durum, renk, eksik_count = "❌ Manuel", "FCEBEB", eksik_count + 1

        degerler = [row-1, f.get('ad',''), f.get('osb',''), f.get('sektor',''),
                    f.get('website',''), email, tel, wa]
        for col, val in enumerate(degerler, 1):
            ws.cell(row=row, column=col, value=val)
        d = ws.cell(row=row, column=9, value=durum)
        d.fill = openpyxl.styles.PatternFill("solid", fgColor=renk)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 4, 40
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    c1, c2, c3, c4 = st.columns(4)
    for c, v, l, renk in [
        (c1, len(firmalar), "Toplam", "#1F4E79"),
        (c2, tam, "✅ Tam", "#1a7a4a"),
        (c3, kismi, "⚠️ Kısmi", "#BA7517"),
        (c4, eksik_count, "❌ Manuel", "#c8401a")
    ]:
        with c:
            st.markdown(f'<div class="stat-box"><div class="stat-val" style="color:{renk}">{v}</div><div class="stat-label">{l}</div></div>', unsafe_allow_html=True)

    st.markdown("")
    st.download_button(
        label=f"⬇️ {st.session_state.sehir}_OSB_Firma_Listesi.xlsx İndir",
        data=buffer,
        file_name=f"{st.session_state.sehir}_OSB_Firma_Listesi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    col_p, col_q = st.columns(2)
    with col_p:
        if st.button("➕ Daha Fazla Firma", use_container_width=True):
            st.session_state.adim = 3
            st.rerun()
    with col_q:
        if st.button("🔄 Yeni Şehir", use_container_width=True):
            for key in ['adim','sehir','osblar','secili_osb','tum_firmalar','gosterilen_ids']:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

    st.balloons()
